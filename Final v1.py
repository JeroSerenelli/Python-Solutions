#______________________ Librerías ______________________#

from importlib.metadata import distribution
import pandas as pd
from openpyxl import Workbook, load_workbook
from datetime import datetime as dt
from openpyxl.utils.cell import column_index_from_string, get_column_letter, coordinate_from_string
import os
import shutil
import pandas as pd
import PySimpleGUI as sg


########______________________ Functions ______________________########
#______________________ Template Creation ______________________#

def CP_form_path():
    sg.theme('NeutralBlue')
    layout = [
        [sg.T("")],
        [sg.Text("Choose the output folder: "),sg.Input(key="-IN1-" ,change_submits=True), sg.FolderBrowse(key="-IN0-")],
        [sg.Button("Exit")]
        ]

    # Building Window

    window = sg.Window('CP File Browser', layout, size= (600,100))

    while True:
        event, values = window.read()
        browsed_path = values["-IN1-"]
        if event == sg.WIN_CLOSED or event == "Exit":
            break
        #elif event == "Ok":
        #    data_path = values["-IN0-"]
        #    break
    return browsed_path

#En vez de usar la función aprovechamos CP_form_path definimos una variable fija 
#directamente para no tener que seleccionar siempre la carpeta

cpform_path = 'C:/Users/JeronimoSerenelli/Box Sync/TA/Reportes TA/Reporte XLS'

def CP_form(formpath): 
    try:
        #Situacion sin errores es que haya 2 o mas elementos en la carpeta
        if "UPDATED TA Balance" in os.listdir(cpform_path)[-2]:
            #print(os.listdir(cpform_path)[-2])
            src = cpform_path + "/{}".format(os.listdir(cpform_path)[-2])
            dst = cpform_path + "/TA Control Points Execution Template.xlsx"
        else:
            src = cpform_path + '/Template/TA Control Points Execution Template.xlsx'
            dst = cpform_path + "/TA Control Points Execution Template.xlsx"
            print("ELSE")
    except IndexError:
        #Situacion en la que hay solo 1 elemento en la carpeta y da error de index
        print("EXCEPT")
        src = cpform_path + '/Template/TA Control Points Execution Template.xlsx'
        dst = cpform_path + "/TA Control Points Execution Template.xlsx"
    return src, dst 

def FileCrea(src, dst):
    shutil.copyfile(src, dst)
    return dst

form_filepath = CP_form(cpform_path)

dst = FileCrea(form_filepath[0],form_filepath[1])


def FileRename(IFSEQ, KFSEQ, DATE):
    src1 = r'C:\Users\JeronimoSerenelli\Box Sync\TA\Reportes TA\Reporte XLS\TA Control Points Execution Template.xlsx'
    file_name = f'C:\\Users\\JeronimoSerenelli\\Box Sync\\TA\\Reportes TA\\Reporte XLS\\FSEQ {IFSEQ} & {KFSEQ} UPDATED TA Balance {DATE}.xlsx'#.format(IFSEQ,KFSEQ,DATE)
    os.rename(src1, file_name)
    return file_name

#______________________ Data Extraction Path ______________________#

def Data_Path():
    sg.theme('NeutralBlue')
    layout = [
        [sg.T("")],
        [sg.Text("Choose a folder: "),sg.Input(key="-IN2-" ,change_submits=True), sg.FolderBrowse(key="-IN-")],
        [sg.Button("Exit")]

        ]

    # Building Window

    window = sg.Window('Data Path Browser', layout, size= (520,100))

    while True:
        event, values = window.read()
        browsed_path = values["-IN2-"]
        if event == sg.WIN_CLOSED or event == "Exit":
            break
        #elif event == "Ok":
        #    data_path = values["-IN-"]
        #    break
    return browsed_path

data_path = Data_Path()
#print(data_path)

#Vamos a saltear por el momento esta función y vamos a dejar fijo el path del file
#data_path = 'C:/Users/JeronimoSerenelli/Box Sync/TA/Reportes TA/2022/03 - Reportes Marzo/15' # Variable temporal

#______________________ File Load ______________________#
# Lo prrimero que deberíamos hacer es pasarnos al directorio donde se encuentra 
# el file para hacer el load o bien cargar el directorio completo del file

wb = load_workbook(dst) 

ws1 = wb["SETTLE Day"]
ws2 = wb["COST"]
ws3 = wb["ACCOUNT_CODE Day"]
ws4 = wb["FNR CHECK"]

def CP_Path(CP):
    for i in os.listdir(data_path):
        if CP in i:
            CP_ABSPATH = data_path + "/{}".format(i)
            control_name = i
            return CP_ABSPATH, control_name

####################################################################
#______________________________ PFAK ______________________________#
####################################################################

#______________________ Workbook Clean ______________________#

pfak_range = ws1['D11':'F14']
for cell in pfak_range:
    for i in cell:
        ws1[i.coordinate] = ""
        
#______________________ CP Report ABS Path ______________________#

#Hacemos un call de la funcion que definimos para buscar el absolute path del Control point

fpath_pfak = CP_Path("PFAK")[0]

#______________________ Data Extract ______________________#  

#Abrimos el archivo y empezamos a tomar parámetros para extraer la información
                                                                    
with open(fpath_pfak) as pfak:
    lines = pfak.readlines()
    counter = 0
    for i in lines:
        if "TOTALS" not in i:
            counter +=1
        else:
            #print("TOTALS is in row", counter)
            break

TOTAL_NET_AMOUNT = float(lines[counter][52:63].replace(",","").replace(" ",""))
#print(TOTAL_NET_AMOUNT)

def batchCatch(row):
    CON_MGR = lines[row][5:8]
    BAT_NUM = lines[row][11:16]
    batch = CON_MGR +" " + BAT_NUM
    return batch

#______________________ CP Check ______________________#
#1st Check

batchs = []
if TOTAL_NET_AMOUNT != 0.00:         # SI TOTAL NET AMMOUT = 0.00 Directamente pasa a otro check
    bcounter = counter
    #print(bcounter)
    for i in range(8,counter-6,6):
            bcounter -= 6
            #print(i)
            if batchCatch(i).isspace() == True:
                pass
            else:
                batchs.append(batchCatch(i))
else:
    print("Move to the next check")

#print(batchs)
#Enhacements & Testing:

# Testear con un file en blanco



#______________________ Data to Excel ______________________#

celdas = []
for col in ws1.iter_cols(min_row=11, max_row=14, min_col=column_index_from_string('D'),max_col=column_index_from_string('F')):
    for cell in col:
        celdas.append(cell.coordinate)
        
#print(celdas)

#counter = 0
#for i in range(len(celdas)):
#    if counter < len(batchs):
#        print(counter)
#        counter=+1
#    else:
#        break

#Completamos las celdas con los batches del txt 

pfak_counter = 0

if len(batchs) >= 1:
    for i in range(len(celdas)):
        if pfak_counter < len(batchs):
            ws1[celdas[i]].value= batchs[i]
            pfak_counter += 1
            #print(ws1[celdas[i]].value)
        else:
            break
            
    #______________________ CP Check ______________________#
    #2nd Check
            
    if len(batchs) > len(celdas):
        excel_last_batch = (coordinate_from_string(celdas[-1]))
        msg_column = column_index_from_string(excel_last_batch[0])+1
        ws1.cell(row = excel_last_batch[1], column = msg_column, value = f"{len(batchs)} Batches - Check TXT")
print("Done PFAK.")


####################################################################
#______________________________ PFAM ______________________________#
####################################################################

#______________________ Workbook Clean ______________________#

pfam_range = ws1['D16':'G19']
for cell in pfam_range:
    for i in cell:
        ws1[i.coordinate] = ""
ws1['H19'] = ""

#______________________ CP Report ABS Path ______________________#

#Hacemos un call de la funcion que definimos para buscar el absolute path del Control point

fpath_pfam = CP_Path("PFAM")[0]
#print(fpath_pfam)

#______________________ Data Extract ______________________#


with open(fpath_pfam) as pfam:
    lines = pfam.readlines()
    counter = 0
    for i in lines:
        if "TOTALS" not in i[1:7]:
            counter +=1
        else:
            #print("TOTALS is in row", counter)
            break

pfam_total_net_amount = float(lines[counter+3][44:61].replace(",","").replace(" ",""))
#print(pfam_total_net_amount)

def batchCatch(row):
    CON_MGR = lines[row][1:5]
    BAT_NUM = lines[row][6:11]
    batch = CON_MGR + BAT_NUM
    return batch

#______________________ CP Check ______________________#
#1st Check
pfam_batchs = []
if pfam_total_net_amount != 0.00:         # SI TOTAL NET AMMOUT = 0.00 Directamente pasa a otro check
    bcounter = counter
    #print(bcounter)
    for i in range(12,counter,5):
            bcounter -= 5
            #print(i)
            if batchCatch(i).isspace() == True:
                pass
            else:
                pfam_batchs.append(batchCatch(i))
    #print(pfam_batchs)
    #print("You have", len(pfam_batchs), "batnums")
else:
    print("Move to the next check")
    

#______________________ Data to Excel ______________________#

#Obtenemos las coordenadas exactas de cada celda sobre las que vamos a iterar introduciendo la data extraída

pfam_celdas = []
for col in ws1.iter_cols(min_row=16, max_row=19, min_col=column_index_from_string('D'),max_col=column_index_from_string('G')):
    for cell in col:
        pfam_celdas.append(cell.coordinate)

#print(pfam_celdas,"\n",len(pfam_celdas))

#BATCH FILL - Iteramos sobre las celdas basándonos en la cantidad de batchs para insertar la info

pfamcounter = 0

if len(pfam_batchs) >= 1:
    for i in range(len(pfam_celdas)):
        if pfamcounter < len(pfam_batchs):
            ws1[pfam_celdas[i]].value= pfam_batchs[i]
            pfamcounter += 1
            #print(ws1[pfam_celdas[i]].value)
        else:
            break
            
#TOTAL NET AMOUNT FILL
ws1.cell(row = 19, column = 8, value = pfam_total_net_amount)

#______________________ CP Check ______________________#
#2nd Check 
#Si la cantidad de batchs es mayor a la cantidad de celdas a imputar lo dejamos detallado en el file
    
if len(pfam_batchs) > len(pfam_celdas): 
    ws1.cell(row = 19, column = 10, value = "Check TXT for additional batches")

print("Done PFAM")

####################################################################
#______________________________ PBEL ______________________________#
####################################################################

#______________________ Workbook Clean ______________________#

pbel_range = ws1['D28':'J29']
for cell in pbel_range:
    for i in cell:
        ws1[i.coordinate] = ""

#______________________ CP Report ABS Path ______________________#

pbel_fpath = CP_Path("PBEL")[0]
#print(pbel_fpath)

#______________________ Data Extract ______________________#
#Definimos las funciones para extraer la info

def dataCatch(row):
    BILL_SRC = lines[row][2:5]
    XMIT_NUM = lines[row][8:12]
    INV_NUM = int(lines[row][38:46].replace(" ",""))
    #if  in lines[row][57:64]:
    UNITS_NUM = lines[row][57:64].replace(" ","")
    RECORDS_NUM = int(lines[row][66:73].replace(" ",""))
    NET_AR = float(lines[row][75:89].replace(" ",""))
    TOTAL_AMOUNTS = float(lines[row][123:137].replace(" ",""))
    batch = BILL_SRC + " " + XMIT_NUM
    data = [batch, INV_NUM, UNITS_NUM, RECORDS_NUM, NET_AR, TOTAL_AMOUNTS]
    return data

def totalsCatch(row):
    INV_NUM = int(lines[row][38:46].replace(" ",""))
    UNITS_NUM = int(lines[row][57:64].replace(" ",""))
    RECORDS_NUM = int(lines[row][66:73].replace(" ",""))
    NET_AR = float(lines[row][75:89].replace(" ",""))
    TOTAL_AMOUNTS = float(lines[row][123:137].replace(" ",""))
    totals_data = [ INV_NUM, UNITS_NUM, RECORDS_NUM, NET_AR, TOTAL_AMOUNTS]
    return totals_data


with open(pbel_fpath) as pbel:
    #global IBM
    #global KYNDRYL
    lines = pbel.readlines()
    counter = 0
    PBEL_IBM = []
    PBEL_KYNDRYL = []
    
    for i in lines:
        #print(counter)
        if '031' in i[2:5]:
            PBEL_IBM.append(dataCatch(counter))
        elif '731' in i[2:5]:
            PBEL_KYNDRYL.append(dataCatch(counter))
        elif "GRAND TOTALS" in i:
            totals = totalsCatch(counter)
        counter += 1
    #print("IBM: ", PBEL_IBM,"\n","\n","KYNDRYL: ", PBEL_KYNDRYL)


#______________________ CP Check ______________________#
#2nd
#Este check lo hacemos ya que la cantidad de unidades de que procesa IBM se suele walcardear

if PBEL_IBM[0][2] == '*******' and PBEL_KYNDRYL != '*******':
    PBEL_IBM[0][2] =  totals[1] - int(PBEL_KYNDRYL[0][2])

#______________________ Data to Excel ______________________#
#Establecemos columnas por las que debemos iterar
pbel_cols = [4,6,7,8,9,10]

#______________________ CHECK 1______________________#
if PBEL_IBM != None:
    for i in pbel_cols:
        ws1.cell(row = 28, column = i, value = PBEL_IBM[0][pbel_cols.index(i)])
if PBEL_KYNDRYL != None:
    for i in pbel_cols:
        ws1.cell(row = 29, column = i, value = PBEL_KYNDRYL[0][pbel_cols.index(i)])

#______________________ CHECK 2______________________#    Check para cuando hay mas de un XMIT num

if PBEL_IBM != None and len(PBEL_IBM) > 1:
    ws1.cell(row= 28, column =column_index_from_string('K'), value = 'Check txt' )
if PBEL_KYNDRYL != None and len(PBEL_KYNDRYL) > 1:
    ws1.cell(row= 28, column =column_index_from_string('K'), value = 'Check txt' )

print("Done PBEL.")

####################################################################
#______________________________ PFAV ______________________________#
####################################################################

#______________________ Workbook Clean ______________________#

pfav_range = ws1['D31':'J32']
for cell in pfav_range:
    for i in cell:
        ws1[i.coordinate] = ""
#______________________ CP Report ABS Path ______________________#

pfav_fpath = CP_Path("PFAV")[0]
#print(pfav_fpath)
#_______________________ Data Extract  _______________________#

def dataCatch(row):
    BILL_SRC = lines[row][20:23].replace(" ","")
    TRANSM_NUM = lines[row][4:9].replace(" ","")
    INV_NUM = int(lines[row+2][41:45].replace(" ","").replace(",",""))
    UNITS_NUM = int(lines[row+2][46:59].replace(" ","").replace(",",""))
    RECORDS_NUM = int(lines[row+2][61:73].replace(" ","").replace(",",""))
    NET_AR = float(lines[row+2][76:91].replace(" ","").replace("$","").replace(",",""))
    CREDITS = float(lines[row+2][116:132].replace(" ","").replace("$","").replace(",",""))
    batch = BILL_SRC + " " + TRANSM_NUM
    data = [batch, INV_NUM, UNITS_NUM, RECORDS_NUM, NET_AR, CREDITS]
    return data

with open(pfav_fpath) as pfav:
    pfav_ibm = []
    pfav_kyndryl= []
    lines = pfav.readlines()
    counter = 0
    for i in lines:
        if '031' in i[20:23]:
            pfav_ibm.append(dataCatch(counter))
        elif '731' in i[20:23]:
            pfav_kyndryl.append(dataCatch(counter))
        counter += 1    
#print("IBM :", pfav_ibm,"\n","\n","KYNDRYL: ", pfav_kyndryl)

#______________________ Data to Excel ______________________#

#Establecemos columnas por las que debemos iterar
pfav_cols = [4,6,7,8,9,10]

#________________________ CP Check ________________________#
if pfav_ibm != None:
    for i in pfav_cols:
        ws1.cell(row = 31, column = i, value = pfav_ibm[0][pfav_cols.index(i)])
        #print(pfav_ibm[0][pfav_cols.index(i)])
if pfav_kyndryl != None:
    for i in pfav_cols:
        ws1.cell(row = 32, column = i, value = pfav_kyndryl[0][pfav_cols.index(i)])
        #print(pfav_kyndryl[0][pfav_cols.index(i)])

#________________________ CP Check ________________________#    Check para cuando hay mas de un XMIT num

if pfav_ibm != None and len(pfav_ibm) > 1:
    ws1.cell(row= 31, column =column_index_from_string('K'), value = 'Check txt')
elif pfav_kyndryl != None and len(pfav_kyndryl) > 1:
    ws1.cell(row= 31, column =column_index_from_string('K'), value = 'Check txt')

print("Done PFAV.")

####################################################################
#______________________________ PFBB ______________________________#
####################################################################

#______________________ Workbook Clean ______________________#

pfbb_range = ws1['F34':'J36']
for cell in pfbb_range:
    for i in cell:
        ws1[i.coordinate] = ""
        
#______________________ CP Report ABS Path ______________________#
        
pfbb_fpath = CP_Path("PFBB")[0]
#print(pfbb_fpath)

#________________________ Data Extract ________________________#

def inputCatch(row):
    INV_NUM = int(lines[row][65:78].replace(" ","").replace(",",""))
    UNITS_NUM = int(lines[row][82:93].replace(" ","").replace(",",""))
    RECORDS_NUM = int(lines[row][53:63].replace(" ","").replace(",",""))
    NET_AR = float(lines[row][29:45].replace(" ","").replace("$","").replace(",",""))
    TOT_CREDITS = float(lines[row][116:133].replace(" ","").replace("$","").replace(",",""))
    #batch = BILL_SRC + " " + TRANSM_NUM
    data = [INV_NUM, UNITS_NUM, RECORDS_NUM, NET_AR, TOT_CREDITS]
    return data

def codedCatch(row):
    
    INV_NUM = int(lines[row][65:78].replace(" ","").replace(",",""))
    UNITS_NUM = int(lines[row][82:93].replace(" ","").replace(",",""))
    RECORDS_NUM = int(lines[row][53:63].replace(" ","").replace(",",""))
    TOT_CREDITS = float(lines[row][116:133].replace(" ","").replace("$","").replace(",",""))
    
    coded_data = [INV_NUM, UNITS_NUM, RECORDS_NUM, TOT_CREDITS]
    return coded_data

with open(pfbb_fpath) as pfbb:
    lines = pfbb.readlines()
    counter = 0
    for i in lines:
        if 'INPUT CTL TOTALS' in i[2:19]:
            INPUT_CTL_TOT = list(inputCatch(counter))
        elif 'TRANS TO BE CODED' in i[2:20]:
            TRANS_TB_CODED = list(codedCatch(counter))
        counter += 1    
#print(INPUT_CTL_TOT,"\n","\n", TRANS_TB_CODED)

#______________________ Data to Excel ______________________#
#Establecemos columnas por las que debemos iterar
ICT_COLS = [6,7,8,9,10]
TTBC_COLS = [6,7,8,10]

#________________________ CP Check ________________________#
#1st
if INPUT_CTL_TOT != None:         #Nunca debería de ser vacío
    for i in ICT_COLS:
        ws1.cell(row = 34, column = i, value = INPUT_CTL_TOT[ICT_COLS.index(i)])

if TRANS_TB_CODED != None:         #Nunca debería de ser vacío
    for i in TTBC_COLS:
        ws1.cell(row = 36, column = i, value = TRANS_TB_CODED[TTBC_COLS.index(i)])

print("Done PFBB.")

####################################################################
#______________________________ PFBD ______________________________#
####################################################################

#______________________ CP Report ABS Path ______________________#
pfbd_fpath = CP_Path("PFBD")[0]
pfbd_rep = CP_Path("PFBD")[1]
#pfbd_fpath
#______________________ Global Variables ______________________#

#Para  poder determinar bien cómo correr el PFBD necesitamos introducir la librería datetime y usar el nombre de 
# los files para determinar la fecha/día y así jugar con las posiciones en las que inputar información

pfbd_cols = [7,8,9,10]

rundays = ["Tuesday","Wednesday","Thursday","Friday","Saturday"]

day ={"Tuesday":['G4', 'H4', 'I4', 'J4'], "Wednesday": ['G5', 'H5', 'I5', 'J5'], 
      "Thursday": ['G6', 'H6', 'I6', 'J6'],"Friday":['G7', 'H7', 'I7', 'J7'],
      "Saturday":['G8', 'H8', 'I8', 'J8']
}

pfbd_range = ws2['D31':'J32']
#______________________ Report Date Extract _____________________# 
#Usamos el nombre del reporte para determinar la fecha en la que se creó y por ende donde debe ir la información.

report_date = str(pfbd_rep[:4]+"-"+pfbd_rep[4:6]+"-"+pfbd_rep[6:8]) 
report_day = dt.fromisoformat(report_date).strftime("%A")

#print(report_date, report_day, dt.now().strftime("%Y%m%d"))
#print("Report day:", report_day)

#______________________ Workbook Clean ______________________#

pfbd_range = ws2['G4':'J8']

if report_day == "Tuesday":
    for cell in pfbd_range:
        for i in cell:
            ws2[i.coordinate] = ""
    else:
        print("Continue")


#______________________ Data Extract ______________________#
def dataCatch_pfbd(row):
    
    ITEMS_NUM = int(lines[row][39:47].replace(" ","").replace(",",""))
    TOTAL_UNITS = int(lines[row][55:68].replace(" ","").replace(",",""))
    TOTAL_DEBITS = float(lines[row][75:94].replace(" ","").replace("$","").replace(",",""))
    TOTAL_CREDITS = float(lines[row][107:121].replace(" ","").replace("$","").replace(",",""))
    
    pfbd_data = [TOTAL_UNITS, ITEMS_NUM , TOTAL_DEBITS, TOTAL_CREDITS]
    return pfbd_data

with open(pfbd_fpath) as pfbd:
    lines = pfbd.readlines()
    counter = 0
    for i in lines:
        if 'VPOF / RMS' in i:
            pfbd_data = dataCatch_pfbd(counter)
        counter += 1    


#print("PFBD Data: ",pfbd_data, "- PFBD Rep name: ",pfbd_rep)

#______________________ Data to Excel ______________________#
#Definimos una función que determina en base al día del reporte, donde debe ir en el excel.

def pfbd_fill(report_day):
    if report_day in rundays:
        #print(report_day)
        #Usamos la información del index para determinar la las coordenadas y con eso inputamos la info en Excel
        
        for i in day[report_day]:
            cl = (column_index_from_string(coordinate_from_string(i)[0]))
            rw = coordinate_from_string(i)[1]
            ws2.cell(row = rw, column = cl, value = pfbd_data[day[report_day].index(i)])
            
            #print(f"Fila: {rw}, Columna: {cl}, Value: {pfbd_data[day[report_day].index(i)]}")

pfbd_fill(report_day)

#________________________ CP Check _________________________# 

#if report_day == "Saturday":
#    for i in Sat:
#        cl = (column_index_from_string(coordinate_from_string(i)[0]))
#        rw = coordinate_from_string(i)[1]
#        ws2.cell(row = rw, column = cl, value = pfbd_data[Sat.index(i)])

print("Done PFBD.")
        

####################################################################
#______________________________ PFBK ______________________________#
####################################################################

#______________________ Workbook Clean _______________________# 

pfbk_range = ['E13','E15','H13','H15','J13','J15']
for cell in pfbk_range:
    ws3[cell] = ""

#______________________ CP Report ABS Path _______________________# 

pfbk_fpath = CP_Path("PFBK")[0]
#print(pfbk_fpath,"\n")

#________________________ Data Extract _________________________# 
# Esta funcion extrae la cantidad de registros y crédito de los invoices procesados y los pendientes
def recs_creds_pfbk(row): 
    
    CODED_NUM_OF_RECS = int(lines[row][38:44].replace(" ","").replace(",",""))+int(lines[row+1][38:44].replace(" ","").replace(",",""))
    CODED_CREDITS = float(lines[row][45:61].replace(" ","").replace("$","").replace(",",""))
    PENDED_NUM_OF_RECS = int(lines[row+2][38:44].replace(" ","").replace(",",""))
    PENDED_CREDITS = float(lines[row+2][45:61].replace(" ","").replace("$","").replace(",",""))
    #return list()
    
    pfbk_data = [CODED_NUM_OF_RECS, CODED_CREDITS, PENDED_NUM_OF_RECS, PENDED_CREDITS]
    return pfbk_data

# Esta funcion extrae el transmital number, el criterio para extraerlo se define en el for loop debajo
def transm_input(row):   
    return lines[row][18:26].replace(" ","")


# Acá abrimos el file y en base al counter vamos definiendo las líneas de donde extraer la informacion en el txt.
with open(pfbk_fpath) as pfbk:
    lines = pfbk.readlines()
    counter = 0
    pfbk_transm_output_coded = []
    pfbk_transm_output_pended = []
    for i in lines:
        if 'CODED' in i[27:36] and 'TOTAL' not in i[18:24]:
            #print(counter)
            pfbk_transm_output_coded.append(transm_input(counter))
        elif 'PENDED' in i[27:34] and 'TOTAL' not in i[18:24]:
            pfbk_transm_output_pended.append(transm_input(counter))
        elif 'TOTAL    CODED' in i[18:33]:
            totals = recs_creds_pfbk(counter)
        counter += 1
#print("Pended Transmital numbers:","\n",pfbk_transm_output_pended)
#print("Coded Transmital numbers:","\n",pfbk_transm_output_coded)
#print("\n", totals)
#________________________ Data Extract _________________________# 

#________________________ Data to Excel _________________________#
# Fill pended Transmital Numbers

tot_pended = ""

for i in pfbk_transm_output_pended:
    tot_pended += str(i) + "  "
ws3.cell(row = 13, column =column_index_from_string('E'), value = tot_pended)

## Fill coded Transmital Numbers
tot_coded = ""

for i in pfbk_transm_output_coded:
    tot_coded += str(i) + "  "
ws3.cell(row = 15, column =column_index_from_string('E'), value = tot_coded)

## Fill pended Record Items & Total Amounts
ws3.cell(row = 13, column =column_index_from_string('H'), value = totals[0])
ws3.cell(row = 13, column =column_index_from_string('J'), value = totals[1])

## Fill Coded Record Items & Total Amounts
ws3.cell(row = 15, column =column_index_from_string('H'), value = totals[2])
ws3.cell(row = 15, column =column_index_from_string('J'), value = totals[3])

print("Done PFBK.")


####################################################################
#______________________________ PFBM ______________________________#
####################################################################

#______________________ Workbook Clean _______________________# 

ws3['D21'] = ""

#______________________ CP Reports ABS Path _______________________#

pfbm_fpath = CP_Path("PFBM")[0]

#______________________ Data Extract _______________________#
# Acá abrimos el file y en base al counter vamos definiendo las líneas de donde extraer la informacion en el txt.

with open(pfbm_fpath) as pfbm:
    lines = pfbm.readlines()
    counter = 0
    transmital_nums = []
    for i in lines:
        if counter > 5 and i[9:16].isspace() == False: 
            if 'TOTAL' not in i[9:16]:
                transmital_nums.append(lines[counter][9:16])
            else:
                break
                #Hace falta poner otro break aca?
        counter +=1
#print(transmital_nums)

#______________________ Data to Excel _______________________#
## Fill pended Transmital Numbers

tot_transmital = ""

for i in transmital_nums:
    tot_transmital += str(i) + "  "
ws3.cell(row = 21, column =column_index_from_string('D'), value = tot_transmital)

print("Done PFBM.")

####################################################################
#______________________________ PFNA ______________________________#
####################################################################

#______________________ Workbook Clean _______________________#

try: 
    ws4.unmerge_cells('D5:D8')
except ValueError:
    pass
    #print("Unmerged.")

pfna_range = ws4['D5':'H8']
for cell in pfna_range:
    for i in cell:
        ws4[i.coordinate] = ""

#______________________ CP Report ABS Path _______________________#
pfna_fpath = CP_Path("PFNA")[0]
#print(pfna_fpath)

#______________________ Data Extract _______________________#
# Acá abrimos el file y en base al counter vamos definiendo las líneas de donde extraer la informacion en el txt.

with open(pfna_fpath) as pfna:
    lines = pfna.readlines()
    counter = 0
    IBM_file = {}
    KYNDRYL_file = {}
    for i in lines:
        if 'FID CONTROL TOTAL:' in i and'031' in i[30:33]:
                IBM_file["ACCTMO"] = lines[counter][42:45].replace(" ","")
                IBM_file["FSN"] =lines[counter][34:38].replace(" ","")
                IBM_file["FID"] = lines[counter][30:33].replace(" ","")
                IBM_file["TOTAL_DEBIT"] = lines[counter][49:67].replace(" ","").replace(",","")
                IBM_file["TOTAL_CREDIT"] = lines[counter][69:87].replace(" ","").replace(",","")
        elif 'FID CONTROL TOTAL:' in i and '731' in i[30:33]:
                KYNDRYL_file["ACCTMO"] = lines[counter][42:45].replace(" ","")
                KYNDRYL_file["FSN"] =lines[counter][34:38].replace(" ","")
                KYNDRYL_file["FID"] = lines[counter][30:33].replace(" ","")              
                KYNDRYL_file["TOTAL_DEBIT"] = lines[counter][49:67].replace(" ","").replace(",","")
                KYNDRYL_file["TOTAL_CREDIT"] = lines[counter][69:87].replace(" ","").replace(",","")
        counter +=1
         
#print(IBM_file, "\n", KYNDRYL_file)

#______________________ Data to Excel _____________________# 
#Global Variables 

pfna_cols = [4,5,6,7,8]

pfna_ibmvalues = list(IBM_file.values())
pfna_kynvalues = list(KYNDRYL_file.values())



counter = 0
for i in pfna_cols:
    for val in pfna_ibmvalues, pfna_kynvalues :
        #print(i, pfna_ibmvalues[counter])
        #print(i, pfna_kynvalues[counter])
        ws4.cell(row= 5, column= i, value = pfna_ibmvalues[counter])
        ws4.cell(row= 6, column= i, value = pfna_kynvalues[counter])
        break
    counter +=1



print("Done PFNA.")

#______________________ PFNA Formatting _____________________#

#replace(" ","")
ws4.merge_cells('D5:D8')



#wb.save("test.xlsx")


####################################################################
#_______________________________ FNR ______________________________#
####################################################################

#______________________ Workbook Clean _______________________#

try: 
    ws4.unmerge_cells('D11:D14')
except ValueError:
    pass
    #print("Unmerged.")

fnr_range = ws4['D11':'H14']
for cell in fnr_range:
    for i in cell:
        ws4[i.coordinate] = ""

#______________________ Workbook Formatting _____________________#

wb.save(dst)
FileRename(IBM_file['FSN'], KYNDRYL_file['FSN'],report_date)